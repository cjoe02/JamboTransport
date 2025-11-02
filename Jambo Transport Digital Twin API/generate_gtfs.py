import csv
import os
from datetime import datetime, timedelta
import zipfile

# Try to import openpyxl, if not available, provide instructions
try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl is not installed. Please run: pip install openpyxl")
    exit(1)

# Configuration
EXCEL_FILE = r"C:\Users\cjose\Desktop\AUT\S2\Mobile\digital-twin-public-transport\majuro-stops.xlsx"
GTFS_DIR = r"C:\Users\cjose\Desktop\AUT\S2\Mobile\digital-twin-public-transport\gtfs"
SPEED_KPH = 15
START_TIME = 6 * 60  # 6:00 AM in minutes
END_TIME = 26 * 60   # 2:00 AM in minutes (26:00 = 02:00 next day)
BREAKS = [(9 * 60, 10), (12 * 60, 10), (15 * 60, 10), (18 * 60, 10)]  # (time in minutes, duration)
BUSES_PER_ROUTE = 2
BUS_OFFSET_MINUTES = 30

def minutes_to_time(minutes):
    """Convert minutes since midnight to HH:MM:SS format"""
    hours = minutes // 60
    mins = minutes % 60
    return f"{hours:02d}:{mins:02d}:00"

def read_excel_data():
    """Read and parse the Excel file"""
    print(f"Reading Excel file: {EXCEL_FILE}")
    wb = load_workbook(EXCEL_FILE)

    # Assuming the first sheet has stops and route data
    sheet = wb.active

    stops = {}
    route_segments = {'A': [], 'B': []}

    # Read the data
    rows = list(sheet.iter_rows(values_only=True))

    print(f"\nAnalyzing Excel structure with {len(rows)} rows...")

    # Parse stops (columns A-D: code, name, lat, long)
    # Starting from row 2 (index 2, after headers at index 1)
    for i in range(2, len(rows)):
        row = rows[i]
        if row[0] and row[1]:  # code and name exist
            try:
                stop_code = str(row[0]).strip()
                stop_name = str(row[1]).strip()
                lat = float(row[2])
                lon = float(row[3])

                stops[stop_name] = {
                    'id': stop_code,
                    'name': stop_name,
                    'lat': lat,
                    'lon': lon
                }
            except (ValueError, TypeError, IndexError):
                pass

    print(f"Found {len(stops)} stops")

    # Build a distance map from the route map data (columns F-H)
    distance_map = {}
    for i in range(2, len(rows)):
        row = rows[i]
        if len(row) > 7 and row[5] and row[6] and row[7]:
            try:
                from_stop = str(row[5]).strip()
                to_stop = str(row[6]).strip()
                distance = float(row[7])
                distance_map[f"{from_stop}|{to_stop}"] = distance
            except (ValueError, TypeError, IndexError):
                pass

    print(f"Built distance map with {len(distance_map)} segments")

    # Parse Route A segments (columns J-K: start location, end location)
    for i in range(2, len(rows)):
        row = rows[i]
        if len(row) > 10 and row[9] and row[10]:  # columns J and K (index 9 and 10)
            try:
                from_stop = str(row[9]).strip()
                to_stop = str(row[10]).strip()

                # Look up distance from distance map
                key = f"{from_stop}|{to_stop}"
                if key in distance_map:
                    route_segments['A'].append({
                        'from': from_stop,
                        'to': to_stop,
                        'distance_km': distance_map[key]
                    })
                else:
                    print(f"Warning: No distance found for Route A segment {from_stop} -> {to_stop}")
            except (ValueError, TypeError, IndexError):
                pass

    # Parse Route B segments (columns M-N: start location, end location)
    for i in range(2, len(rows)):
        row = rows[i]
        if len(row) > 13 and row[12] and row[13]:  # columns M and N (index 12 and 13)
            try:
                from_stop = str(row[12]).strip()
                to_stop = str(row[13]).strip()

                # Look up distance from distance map
                key = f"{from_stop}|{to_stop}"
                if key in distance_map:
                    route_segments['B'].append({
                        'from': from_stop,
                        'to': to_stop,
                        'distance_km': distance_map[key]
                    })
                else:
                    print(f"Warning: No distance found for Route B segment {from_stop} -> {to_stop}")
            except (ValueError, TypeError, IndexError):
                pass

    return stops, route_segments

def calculate_trip_times(segments, stops):
    """Calculate arrival times for each stop based on distance and speed"""
    stop_times = []
    cumulative_time = 0  # in minutes

    # Add first stop
    if segments:
        first_stop = segments[0]['from']
        if first_stop in stops:
            stop_times.append({
                'stop_id': stops[first_stop]['id'],
                'stop_name': first_stop,
                'time_offset': 0
            })

    # Calculate times for subsequent stops
    for segment in segments:
        travel_time = (segment['distance_km'] / SPEED_KPH) * 60  # convert to minutes
        cumulative_time += travel_time

        to_stop = segment['to']
        if to_stop in stops:
            stop_times.append({
                'stop_id': stops[to_stop]['id'],
                'stop_name': to_stop,
                'time_offset': cumulative_time
            })

    return stop_times

def generate_trips_for_route(route_id, stop_times, start_offset=0):
    """Generate multiple trips throughout the day for a route"""
    trips = []
    current_time = START_TIME + start_offset
    trip_number = 1

    # Calculate total trip duration
    if stop_times:
        trip_duration = stop_times[-1]['time_offset']
    else:
        trip_duration = 0

    while current_time + trip_duration <= END_TIME:
        # Check if we need to add a break
        break_added = False
        for break_time, break_duration in BREAKS:
            if current_time <= break_time < current_time + trip_duration:
                # Schedule break after this trip
                current_time += trip_duration
                if current_time <= break_time < current_time + 30:  # If break is within 30 min after trip
                    current_time = break_time + break_duration
                    break_added = True
                break

        if not break_added:
            trips.append({
                'trip_id': f"{route_id}_TRIP{trip_number:03d}",
                'start_time': current_time,
                'stop_times': stop_times
            })
            current_time += trip_duration
            trip_number += 1

        # Small buffer between trips
        current_time += 5  # 5 minute buffer

    return trips

def create_gtfs_files(stops, route_segments):
    """Create all GTFS CSV files"""

    # Create GTFS directory
    os.makedirs(GTFS_DIR, exist_ok=True)
    print(f"\nCreating GTFS files in: {GTFS_DIR}")

    # 1. Create stops.txt
    with open(os.path.join(GTFS_DIR, 'stops.txt'), 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['stop_id', 'stop_name', 'stop_lat', 'stop_lon'])
        for stop in stops.values():
            writer.writerow([stop['id'], stop['name'], stop['lat'], stop['lon']])

    print(f"Created stops.txt with {len(stops)} stops")

    # 2. Create routes.txt
    with open(os.path.join(GTFS_DIR, 'routes.txt'), 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['route_id', 'route_short_name', 'route_long_name', 'route_type'])
        writer.writerow(['ROUTE_A', 'A', 'Route A', '3'])
        writer.writerow(['ROUTE_B', 'B', 'Route B', '3'])

    print("Created routes.txt with 2 routes")

    # 3. Calculate stop times for each route
    route_stop_times = {}
    route_distances = {}

    for route_key in ['A', 'B']:
        segments = route_segments[route_key]
        stop_times = calculate_trip_times(segments, stops)
        route_stop_times[route_key] = stop_times
        route_distances[route_key] = sum(seg['distance_km'] for seg in segments)

    # 4. Generate trips for each route (2 buses per route)
    all_trips = []
    trips_per_route = {'A': 0, 'B': 0}

    for route_key in ['A', 'B']:
        route_id = f"ROUTE_{route_key}"
        stop_times = route_stop_times[route_key]

        # For Route B, create reverse stop times
        reverse_stop_times = None
        if route_key == 'B':
            # Reverse the stop sequence for return trip
            reverse_segments = list(reversed(route_segments[route_key]))
            # Create new stop times with proper time offsets
            reverse_stop_times = []
            cumulative_time = 0

            # Add first stop (which was the last stop in forward direction)
            reverse_stop_times.append({
                'stop_id': stop_times[-1]['stop_id'],
                'stop_name': stop_times[-1]['stop_name'],
                'time_offset': 0
            })

            # Calculate times for reverse direction
            for i, segment in enumerate(reverse_segments):
                travel_time = (segment['distance_km'] / SPEED_KPH) * 60
                cumulative_time += travel_time

                # Find the stop that corresponds to 'from' in reverse (was 'to' in forward)
                stop_name = segment['from']
                matching_stop = next((st for st in stop_times if st['stop_name'] == stop_name), None)
                if matching_stop:
                    reverse_stop_times.append({
                        'stop_id': matching_stop['stop_id'],
                        'stop_name': stop_name,
                        'time_offset': cumulative_time
                    })

        for bus_num in range(BUSES_PER_ROUTE):
            offset = bus_num * BUS_OFFSET_MINUTES
            trips = generate_trips_for_route(route_id, stop_times, offset)

            # Add bus identifier to trip_id and direction
            trip_counter = 1
            for trip in trips:
                # Forward trip
                trip['trip_id'] = f"{route_id}_BUS{bus_num+1}_TRIP{trip_counter:03d}"
                trip['route_id'] = route_id
                trip['direction_id'] = 0
                all_trips.append(trip)

                # For Route B, add reverse trip immediately after
                if route_key == 'B' and reverse_stop_times:
                    reverse_trip = {
                        'trip_id': f"{route_id}_BUS{bus_num+1}_TRIP{trip_counter:03d}_RETURN",
                        'start_time': trip['start_time'] + stop_times[-1]['time_offset'] + 5,  # Start 5 min after arrival
                        'stop_times': reverse_stop_times,
                        'route_id': route_id,
                        'direction_id': 1
                    }
                    all_trips.append(reverse_trip)
                    trips_per_route[route_key] += 1

                trip_counter += 1

            trips_per_route[route_key] += len(trips)

    # 5. Create trips.txt
    with open(os.path.join(GTFS_DIR, 'trips.txt'), 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['route_id', 'service_id', 'trip_id', 'trip_headsign', 'direction_id'])
        for trip in all_trips:
            route_key = trip['route_id'].split('_')[1]
            last_stop = trip['stop_times'][-1]['stop_name'] if trip['stop_times'] else ''
            direction_id = trip.get('direction_id', 0)
            writer.writerow([trip['route_id'], 'WEEKDAY', trip['trip_id'], last_stop, direction_id])

    print(f"Created trips.txt with {len(all_trips)} trips")

    # 6. Create stop_times.txt
    with open(os.path.join(GTFS_DIR, 'stop_times.txt'), 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['trip_id', 'arrival_time', 'departure_time', 'stop_id', 'stop_sequence'])

        for trip in all_trips:
            for seq, stop_time in enumerate(trip['stop_times'], 1):
                arrival_minutes = int(trip['start_time'] + stop_time['time_offset'])
                arrival_time = minutes_to_time(arrival_minutes)
                departure_time = arrival_time  # Same as arrival for simplicity

                writer.writerow([
                    trip['trip_id'],
                    arrival_time,
                    departure_time,
                    stop_time['stop_id'],
                    seq
                ])

    print("Created stop_times.txt")

    # 7. Create calendar.txt
    with open(os.path.join(GTFS_DIR, 'calendar.txt'), 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['service_id', 'monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday', 'start_date', 'end_date'])
        # Service runs all days
        writer.writerow(['WEEKDAY', '1', '1', '1', '1', '1', '1', '1', '20250101', '20251231'])

    print("Created calendar.txt")

    # Return summary data
    return {
        'stops_count': len(stops),
        'route_a_segments': len(route_segments['A']),
        'route_b_segments': len(route_segments['B']),
        'route_a_distance': route_distances['A'],
        'route_b_distance': route_distances['B'],
        'route_a_trips': trips_per_route['A'],
        'route_b_trips': trips_per_route['B']
    }

def main():
    print("="*60)
    print("GTFS File Generator")
    print("="*60)

    # Read Excel data
    stops, route_segments = read_excel_data()

    if not stops:
        print("\nError: No stops found in Excel file!")
        return

    print(f"\nFound {len(stops)} stops")
    print(f"Found {len(route_segments['A'])} segments for Route A")
    print(f"Found {len(route_segments['B'])} segments for Route B")

    # Create GTFS files
    summary = create_gtfs_files(stops, route_segments)

    # Print summary
    print("\n" + "="*60)
    print("GENERATION COMPLETE - SUMMARY")
    print("="*60)
    print(f"1. Total stops: {summary['stops_count']}")
    print(f"\n2. Route segments:")
    print(f"   - Route A: {summary['route_a_segments']} segments")
    print(f"   - Route B: {summary['route_b_segments']} segments")
    print(f"\n3. Total distance per route:")
    print(f"   - Route A: {summary['route_a_distance']:.2f} km")
    print(f"   - Route B: {summary['route_b_distance']:.2f} km")
    print(f"\n4. Trips generated per route:")
    print(f"   - Route A: {summary['route_a_trips']} trips ({BUSES_PER_ROUTE} buses)")
    print(f"   - Route B: {summary['route_b_trips']} trips ({BUSES_PER_ROUTE} buses)")
    print(f"\nAll GTFS files created in: {GTFS_DIR}")
    print("="*60)

if __name__ == "__main__":
    main()
